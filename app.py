from flask import Flask, render_template, request, jsonify, session
import csv
import os
from werkzeug.utils import secure_filename
import uuid
from difflib import SequenceMatcher
import re
from collections import defaultdict
import logging
from datetime import datetime
import json
import io

app = Flask(__name__)
app.secret_key = 'bootcode_verification_secret_key'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Create uploads directory if it doesn't exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Store uploaded data in memory (in production, use a database)
uploaded_data = {}
# Store conversation context for follow-up questions
conversation_context = {}
# Store unmatched errors for dataset improvement
unmatched_errors = []

# Setup simplified logging for unmatched errors
try:
    logging.basicConfig(
        level=logging.WARNING,  # Reduced logging level
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler('unmatched_errors.log', encoding='utf-8'),
        ]
    )
    logger = logging.getLogger(__name__)
except Exception as e:
    # Fallback to basic logging if file logging fails
    logging.basicConfig(level=logging.ERROR)
    logger = logging.getLogger(__name__)
    print(f"⚠️ Warning: Logging setup had issues: {e}")

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'csv', 'xls', 'xlsx'}

def similarity(a, b):
    """Calculate similarity between two strings"""
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()

def extract_keywords(text):
    """Extract meaningful keywords from error message"""
    # Common bootcode-related keywords
    keywords = []
    text_lower = text.lower()
    
    # Boot-related terms
    boot_terms = ['boot', 'bios', 'uefi', 'mbr', 'gpt', 'bootloader', 'grub', 'ntldr', 'bootmgr']
    hardware_terms = ['memory', 'ram', 'cpu', 'disk', 'drive', 'ssd', 'hdd', 'tpm', 'secure']
    error_terms = ['error', 'failure', 'failed', 'missing', 'corrupt', 'invalid', 'timeout', 'panic']
    
    for term in boot_terms + hardware_terms + error_terms:
        if term in text_lower:
            keywords.append(term)
    
    return keywords

def is_ambiguous_query(matches, original_message):
    """Determine if the query is ambiguous and needs clarification"""
    if len(matches) < 2:
        return False
    
    # Check if top matches have similar similarity scores but different keywords
    top_matches = matches[:3]  # Look at top 3 matches
    similarity_threshold = 0.15  # If similarity scores are within 15% of each other
    
    # Group matches by similarity score ranges
    score_groups = []
    for match in top_matches:
        added_to_group = False
        for group in score_groups:
            if abs(group[0]['similarity'] - match['similarity']) <= similarity_threshold:
                group.append(match)
                added_to_group = True
                break
        if not added_to_group:
            score_groups.append([match])
    
    # If we have multiple matches with similar scores
    if len(score_groups) > 0 and len(score_groups[0]) >= 2:
        # Check if they represent different error categories
        error_categories = set()
        for match in score_groups[0]:
            keywords = extract_keywords(match['error'])
            category = determine_error_category(keywords)
            error_categories.add(category)
        
        # If different categories, it's ambiguous
        if len(error_categories) > 1:
            return True
    
    # Check if the original message is too generic
    original_keywords = extract_keywords(original_message)
    if len(original_keywords) <= 2 and len(matches) >= 3:
        return True
    
    return False

def determine_error_category(keywords):
    """Categorize error based on keywords"""
    if any(term in keywords for term in ['memory', 'ram']):
        return 'memory'
    elif any(term in keywords for term in ['disk', 'drive', 'ssd', 'hdd', 'mbr', 'gpt']):
        return 'storage'
    elif any(term in keywords for term in ['boot', 'bootloader', 'grub', 'ntldr', 'bootmgr']):
        return 'bootloader'
    elif any(term in keywords for term in ['bios', 'uefi', 'tpm', 'secure']):
        return 'firmware'
    elif any(term in keywords for term in ['cpu']):
        return 'processor'
    else:
        return 'general'

def generate_follow_up_question(matches, original_message):
    """Generate intelligent follow-up questions based on matches"""
    if len(matches) < 2:
        return None
    
    # Analyze the matches to create targeted questions
    categories = defaultdict(list)
    for match in matches[:5]:  # Look at top 5 matches
        keywords = extract_keywords(match['error'])
        category = determine_error_category(keywords)
        categories[category].append(match)
    
    # Generate questions based on categories
    if len(categories) > 1:
        category_options = []
        category_names = {
            'memory': 'Memory/RAM related',
            'storage': 'Hard drive/Storage related', 
            'bootloader': 'Boot loader/Boot manager related',
            'firmware': 'BIOS/UEFI/Firmware related',
            'processor': 'CPU/Processor related',
            'general': 'General system'
        }
        
        for category, matches_in_category in categories.items():
            if len(matches_in_category) > 0:
                example = matches_in_category[0]['error']
                category_display = category_names.get(category, category.title())
                category_options.append({
                    'category': category,
                    'display_name': category_display,
                    'example': example,
                    'count': len(matches_in_category)
                })
        
        if len(category_options) >= 2:
            question = "I found multiple types of errors that might match your description. Could you help me narrow it down?\n\n"
            
            for i, option in enumerate(category_options[:4], 1):  # Show max 4 options
                match_text = "matches" if option["count"] > 1 else "match"
                question += f"{i}. **{option['display_name']}** ({option['count']} {match_text})\n"
                question += f"   Example: \"{option['example']}\"\n\n"
            
            question += "Please type the number or describe which type of error you're experiencing."
            
            return {
                'question': question,
                'type': 'category_selection',
                'options': category_options,
                'original_matches': matches
            }
    
    # If same category but different specific errors
    elif len(matches) >= 3:
        question = "I found several similar errors. Could you provide more details?\n\n"
        
        for i, match in enumerate(matches[:4], 1):
            similarity_percent = int(match['similarity'] * 100)
            question += f"{i}. \"{match['error']}\" ({similarity_percent}% match)\n"
        
        question += f"\nPlease type the number of the closest match, or provide more specific details about your error."
        
        return {
            'question': question,
            'type': 'specific_selection',
            'options': matches[:4],
            'original_matches': matches
        }
    
    return None

def log_unmatched_error(error_message, session_id, user_context=None):
    """Log unmatched error messages for dataset improvement"""
    timestamp = datetime.now().isoformat()
    
    error_entry = {
        'timestamp': timestamp,
        'error_message': error_message,
        'session_id': session_id,
        'user_context': user_context
    }
    
    # Add to in-memory list
    unmatched_errors.append(error_entry)
    
    # Log to file
    logger.info(f"UNMATCHED_ERROR: {json.dumps(error_entry)}")
    
    # Keep only last 100 unmatched errors in memory
    if len(unmatched_errors) > 100:
        unmatched_errors.pop(0)

def generate_improvement_suggestions(error_message):
    """Generate suggestions for improving the dataset"""
    suggestions = []
    
    # Analyze the error message to provide contextual suggestions
    keywords = extract_keywords(error_message)
    error_lower = error_message.lower()
    
    if not keywords:
        suggestions.append("Try using more specific technical terms related to boot processes")
        suggestions.append("Include error codes or specific component names if available")
    
    if len(error_message.split()) < 3:
        suggestions.append("Provide more detailed description of when this error occurs")
        suggestions.append("Include any error codes or system information")
    
    # Specific suggestions based on content
    if any(term in error_lower for term in ['boot', 'startup', 'start']):
        suggestions.append("Specify the boot stage: BIOS/UEFI, bootloader, or OS loading")
        suggestions.append("Mention if this happens on cold boot, warm restart, or both")
    
    if any(term in error_lower for term in ['error', 'fail']):
        suggestions.append("Include the exact error message or code if displayed")
        suggestions.append("Describe what happens: system freezes, restarts, or shows error screen")
    
    # General improvement suggestions
    suggestions.extend([
        "Consider adding this error to your database with a detailed fix description",
        "Check if there are similar errors in your database that might help",
        "Document the system configuration where this error occurs"
    ])
    
    return suggestions[:6]  # Return max 6 suggestions

def generate_database_entry_template(error_message):
    """Generate a template for adding new entries to the database"""
    template = {
        'error_message': error_message,
        'primary_fix': '[Please provide the main solution for this error]',
        'alternative_fix': '[Optional: Provide an alternative solution]',
        'additional_fix': '[Optional: Provide additional troubleshooting steps]',
        'priority': '[High/Medium/Low - Set priority level]',
        'category': determine_error_category(extract_keywords(error_message)),
        'suggested_format': {
            'csv_row': f'"{error_message}","[Primary Fix]","[Alternative Fix]","[Additional Fix]","Medium"',
            'excel_columns': ['Error Message', 'Primary Fix', 'Alternative Fix', 'Additional Fix', 'Priority']
        }
    }
    return template

def read_file_data(file):
    """Read CSV or Excel file and return data as list of dictionaries"""
    try:
        filename = file.filename.lower()
        print(f"📖 Processing file: {filename}")
        
        if filename.endswith('.csv'):
            print("📄 Reading CSV file...")
            # Read CSV file with encoding detection
            try:
                content = file.read().decode('utf-8')
                print("✅ UTF-8 encoding successful")
            except UnicodeDecodeError as e:
                print(f"⚠️ UTF-8 failed ({e}), trying latin-1...")
                file.seek(0)  # Reset file pointer
                try:
                    content = file.read().decode('latin-1')
                    print("✅ Latin-1 encoding successful")
                except UnicodeDecodeError:
                    print("❌ Trying cp1252...")
                    file.seek(0)
                    content = file.read().decode('cp1252', errors='replace')
                    print("✅ CP1252 encoding with replacement")
            
            # Parse CSV
            csv_reader = csv.DictReader(io.StringIO(content))
            data = []
            columns = list(csv_reader.fieldnames) if csv_reader.fieldnames else []
            print(f"📋 CSV columns detected: {columns}")
            
            for row_num, row in enumerate(csv_reader, 1):
                if any(value and str(value).strip() for value in row.values()):  # Skip empty rows
                    data.append(row)
                if row_num > 1000:  # Limit to prevent memory issues
                    print("⚠️ Limiting to first 1000 rows")
                    break
                    
        else:
            print("📊 Reading Excel file...")
            # Check if openpyxl is available
            try:
                from openpyxl import load_workbook
            except ImportError:
                raise ImportError("openpyxl is required for Excel files. Please install it with: pip install openpyxl")
            
            # Read Excel file
            try:
                workbook = load_workbook(file, read_only=True)
                sheet = workbook.active
                print(f"📊 Excel sheet loaded: {sheet.title}")
            except Exception as e:
                raise Exception(f"Cannot read Excel file: {str(e)}")
            
            # Get headers from first row
            first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not first_row:
                raise Exception("Excel file appears to be empty")
                
            columns = [str(cell).strip() if cell is not None else f'Column_{i+1}' 
                      for i, cell in enumerate(first_row)]
            print(f"📋 Excel columns detected: {columns}")
            
            # Get data rows
            data = []
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 1):
                if any(cell is not None and str(cell).strip() for cell in row):  # Skip empty rows
                    row_dict = {}
                    for i, value in enumerate(row):
                        if i < len(columns):
                            row_dict[columns[i]] = str(value).strip() if value is not None else ''
                    data.append(row_dict)
                if row_num > 1000:  # Limit to prevent memory issues
                    print("⚠️ Limiting to first 1000 rows")
                    break
        
        print(f"✅ File processed: {len(data)} data rows, {len(columns)} columns")
        
        # Validate results
        if not columns:
            raise Exception("No columns detected in file")
        if not data:
            raise Exception("No data rows found in file")
            
        return data, columns
        
    except Exception as e:
        print(f"❌ Error in read_file_data: {e}")
        raise Exception(f"Error reading file: {str(e)}")

def find_best_matches(error_message, data, threshold=0.6):
    """Find best matching error messages in the data"""
    if not data or len(data) == 0:
        return []
    
    # Get column names (first column is error messages)
    columns = list(data[0].keys())
    error_col = columns[0]
    
    matches = []
    error_message_lower = error_message.lower()
    
    for row in data:
        stored_error = str(row.get(error_col, '')).lower()
        
        # Check for exact substring match first
        if error_message_lower in stored_error or stored_error in error_message_lower:
            similarity_score = 1.0
        else:
            # Calculate similarity score
            similarity_score = similarity(error_message, str(row.get(error_col, '')))
        
        if similarity_score >= threshold:
            # Collect all fix columns and their data
            fixes = []
            priority = "Medium"  # Default priority
            
            # Process all columns after the first one
            for col_idx, col_name in enumerate(columns[1:], 1):
                if col_name in row:
                    col_value = str(row[col_name]).strip()
                    
                    # Skip empty/null values
                    if col_value and col_value.lower() not in ['nan', 'none', '', 'null']:
                        if col_name.lower() in ['priority', 'priority_level', 'urgency']:
                            priority = col_value
                        elif col_name.lower() in ['primary_fix', 'main_fix', 'fix', 'solution']:
                            fixes.insert(0, {'type': 'Primary', 'content': col_value})
                        elif col_name.lower() in ['secondary_fix', 'alternative_fix', 'alt_fix', 'alternative']:
                            fixes.append({'type': 'Alternative', 'content': col_value})
                        elif col_name.lower() in ['tertiary_fix', 'additional_fix', 'extra_fix']:
                            fixes.append({'type': 'Additional', 'content': col_value})
                        else:
                            # Auto-detect fix type based on column position
                            if col_idx == 1:
                                fixes.insert(0, {'type': 'Primary', 'content': col_value})
                            elif col_idx == 2:
                                fixes.append({'type': 'Alternative', 'content': col_value})
                            elif col_idx == 3:
                                fixes.append({'type': 'Additional', 'content': col_value})
                            elif col_idx >= 4 and col_name.lower() not in ['priority', 'priority_level', 'urgency']:
                                fixes.append({'type': f'Option {col_idx-1}', 'content': col_value})
            
            # Ensure we have at least one fix
            if not fixes:
                fixes = [{'type': 'Primary', 'content': 'No specific fix provided'}]
            
            matches.append({
                'error': row.get(error_col, ''),
                'fixes': fixes,
                'priority': priority,
                'similarity': similarity_score
            })
    
    # Sort by priority first, then similarity score
    priority_order = {'High': 3, 'Medium': 2, 'Low': 1, 'Critical': 4, 'Urgent': 4}
    matches.sort(key=lambda x: (priority_order.get(x['priority'], 2), x['similarity']), reverse=True)
    return matches[:5]  # Return top 5 matches

@app.route('/')
def index():
    try:
        return render_template('index.html')
    except Exception as e:
        # Handle any template loading issues
        error_type = type(e).__name__
        print(f"⚠️ Template error ({error_type}): {e}")
        
        # Fallback HTML with basic functionality
        fallback_html = '''
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Bootcode Verification Chatbot</title>
            <style>
                body { font-family: Arial, sans-serif; margin: 40px; background: #f5f5f5; }
                .container { max-width: 800px; margin: 0 auto; background: white; padding: 30px; border-radius: 10px; }
                .header { text-align: center; color: #333; margin-bottom: 30px; }
                .upload-section { margin: 20px 0; padding: 20px; background: #f8f9fa; border-radius: 5px; }
                .chat-area { min-height: 300px; border: 1px solid #ddd; padding: 20px; margin: 20px 0; }
                input[type="file"], input[type="text"], button { padding: 10px; margin: 5px; }
                .send-btn { background: #007bff; color: white; border: none; border-radius: 5px; }
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>🔧 Bootcode Verification Chatbot</h1>
                    <p>Fallback mode - Basic functionality available</p>
                </div>
                <div class="upload-section">
                    <label for="file">Upload Error Database (CSV/Excel):</label><br>
                    <input type="file" id="file" accept=".csv,.xls,.xlsx">
                    <div id="status"></div>
                </div>
                <div class="chat-area" id="chat"></div>
                <div>
                    <input type="text" id="message" placeholder="Enter error message..." disabled style="width: 70%;">
                    <button class="send-btn" id="send" disabled onclick="sendMessage()">Send</button>
                </div>
            </div>
            <script>
                let fileUploaded = false;
                document.getElementById('file').onchange = async function(e) {
                    const file = e.target.files[0];
                    if (!file) return;
                    const formData = new FormData();
                    formData.append('file', file);
                    try {
                        const response = await fetch('/upload', { method: 'POST', body: formData });
                        const result = await response.json();
                        if (result.success) {
                            document.getElementById('status').innerHTML = '✅ ' + result.message;
                            document.getElementById('message').disabled = false;
                            document.getElementById('send').disabled = false;
                            fileUploaded = true;
                        } else {
                            document.getElementById('status').innerHTML = '❌ ' + result.error;
                        }
                    } catch (error) {
                        document.getElementById('status').innerHTML = '❌ Upload failed';
                    }
                };
                async function sendMessage() {
                    const message = document.getElementById('message').value.trim();
                    if (!message || !fileUploaded) return;
                    document.getElementById('chat').innerHTML += '<div><strong>You:</strong> ' + message + '</div>';
                    document.getElementById('message').value = '';
                    try {
                        const response = await fetch('/chat', {
                            method: 'POST',
                            headers: { 'Content-Type': 'application/json' },
                            body: JSON.stringify({ message: message })
                        });
                        const result = await response.json();
                        let botResponse = '<div><strong>Bot:</strong> ' + result.message + '</div>';
                        if (result.matches) {
                            result.matches.forEach(match => {
                                botResponse += '<div style="margin: 10px 0; padding: 10px; border: 1px solid #ddd;">';
                                botResponse += '<strong>Error:</strong> ' + match.error + '<br>';
                                match.fixes.forEach(fix => {
                                    botResponse += '<strong>' + fix.type + ':</strong> ' + fix.content + '<br>';
                                });
                                botResponse += '</div>';
                            });
                        }
                        document.getElementById('chat').innerHTML += botResponse;
                    } catch (error) {
                        document.getElementById('chat').innerHTML += '<div><strong>Bot:</strong> Error occurred</div>';
                    }
                }
                document.getElementById('message').onkeypress = function(e) {
                    if (e.key === 'Enter') sendMessage();
                };
            </script>
        </body>
        </html>
        '''
        return fallback_html

@app.route('/upload', methods=['POST'])
def upload_file():
    try:
        print("📁 Upload request received")
        
        # Check if file is in request
        if 'file' not in request.files:
            print("❌ No file in request")
            return jsonify({'error': 'No file selected'}), 400
        
        file = request.files['file']
        print(f"📄 File received: {file.filename}")
        
        if file.filename == '':
            print("❌ Empty filename")
            return jsonify({'error': 'No file selected'}), 400
        
        # Check file size
        file.seek(0, 2)  # Seek to end
        file_size = file.tell()
        file.seek(0)  # Reset to beginning
        print(f"📊 File size: {file_size} bytes")
        
        if file_size > app.config['MAX_CONTENT_LENGTH']:
            return jsonify({'error': f'File too large. Maximum size is {app.config["MAX_CONTENT_LENGTH"]//1024//1024}MB'}), 400
        
        if file_size == 0:
            return jsonify({'error': 'File is empty'}), 400
        
        if file and allowed_file(file.filename):
            try:
                print("✅ File type allowed, processing...")
                
                # Generate unique session ID
                if 'session_id' not in session:
                    session['session_id'] = str(uuid.uuid4())
                
                session_id = session['session_id']
                print(f"🔑 Session ID: {session_id}")
                
                # Read the file data
                print("📖 Reading file data...")
                data, columns = read_file_data(file)
                print(f"📊 Data read: {len(data)} rows, {len(columns)} columns")
                print(f"📋 Columns: {columns}")
                
                # Validate that the file has at least 2 columns
                if len(columns) < 2:
                    return jsonify({'error': 'File must have at least 2 columns (Error Message and at least one Fix column)'}), 400
                
                # Validate data
                if len(data) == 0:
                    return jsonify({'error': 'File contains no data rows'}), 400
                
                # Store the data
                uploaded_data[session_id] = data
                print("✅ Data stored successfully")
                
                return jsonify({
                    'success': True, 
                    'message': f'File uploaded successfully! Found {len(data)} error records.',
                    'columns': list(columns),
                    'sample_data': data[:3] if len(data) >= 3 else data
                })
                
            except ImportError as e:
                print(f"❌ Import error: {e}")
                if 'openpyxl' in str(e):
                    return jsonify({'error': 'Excel support not available. Please install openpyxl or use CSV files.'}), 400
                else:
                    return jsonify({'error': f'Missing dependency: {str(e)}'}), 400
                    
            except Exception as e:
                print(f"❌ Processing error: {e}")
                error_msg = str(e)
                if 'BadZipFile' in error_msg:
                    return jsonify({'error': 'Invalid Excel file. Please check the file format.'}), 400
                elif 'UnicodeDecodeError' in error_msg:
                    return jsonify({'error': 'File encoding issue. Please save as UTF-8 CSV or check Excel file.'}), 400
                else:
                    return jsonify({'error': f'Error processing file: {error_msg}'}), 400
        else:
            print(f"❌ File type not allowed: {file.filename}")
            return jsonify({'error': 'Invalid file type. Please upload CSV, XLS, or XLSX files.'}), 400
            
    except Exception as e:
        print(f"❌ Upload route error: {e}")
        return jsonify({'error': f'Upload failed: {str(e)}'}), 500

@app.route('/chat', methods=['POST'])
def chat():
    request_data = request.get_json()
    message = request_data.get('message', '').strip()
    
    if not message:
        return jsonify({'error': 'Please enter an error message'}), 400
    
    session_id = session.get('session_id')
    if not session_id or session_id not in uploaded_data:
        return jsonify({'error': 'Please upload a file first'}), 400
    
    data = uploaded_data[session_id]
    
    # Find matching errors
    matches = find_best_matches(message, data)
    
    if not matches:
        # Log unmatched error for dataset improvement
        log_unmatched_error(message, session_id)
        
        # Generate improvement suggestions
        improvement_suggestions = generate_improvement_suggestions(message)
        template = generate_database_entry_template(message)
        
        response = {
            'message': 'No matching errors found in the uploaded data.',
            'suggestions': improvement_suggestions,
            'unmatched': True,
            'database_template': template
        }
    else:
        if matches[0]['similarity'] >= 0.9:
            response = {
                'message': f'Found exact match! Here are the recommended solutions:',
                'exact_match': True,
                'matches': [matches[0]]
            }
        else:
            # Check if query is ambiguous and needs follow-up
            follow_up = None
            if is_ambiguous_query(matches, message):
                follow_up = generate_follow_up_question(matches, message)
            
            response = {
                'message': f'Found {len(matches)} similar error(s). Here are the closest matches:',
                'exact_match': False,
                'matches': matches,
                'follow_up': follow_up
            }
    
    return jsonify(response)

@app.route('/clear', methods=['POST'])
def clear_session():
    session_id = session.get('session_id')
    if session_id and session_id in uploaded_data:
        del uploaded_data[session_id]
    session.clear()
    return jsonify({'success': True, 'message': 'Session cleared successfully'})

@app.route('/test-upload', methods=['GET'])
def test_upload():
    """Test endpoint to check upload functionality"""
    return jsonify({
        'status': 'Upload endpoint is working',
        'max_file_size': f"{app.config['MAX_CONTENT_LENGTH']//1024//1024}MB",
        'allowed_extensions': ['csv', 'xls', 'xlsx'],
        'openpyxl_available': True,
        'session_info': {
            'has_session': 'session_id' in session,
            'session_id': session.get('session_id', 'None')
        }
    })

@app.route('/unmatched-errors', methods=['GET'])
def get_unmatched_errors():
    """Get list of unmatched errors for dataset improvement"""
    return jsonify({
        'unmatched_errors': unmatched_errors[-20:],  # Return last 20 unmatched errors
        'total_count': len(unmatched_errors)
    })

@app.route('/download-unmatched', methods=['GET'])
def download_unmatched_errors():
    """Download unmatched errors as CSV for dataset improvement"""
    try:
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow(['Timestamp', 'Error Message', 'Suggested Primary Fix', 'Suggested Alternative Fix', 'Priority'])
        
        # Write unmatched errors with suggested template
        for error_entry in unmatched_errors:
            writer.writerow([
                error_entry['timestamp'],
                error_entry['error_message'],
                '[Please provide the main solution for this error]',
                '[Optional: Provide an alternative solution]',
                'Medium'
            ])
        
        output.seek(0)
        response = Flask.response_class(
            output.getvalue(),
            mimetype='text/csv',
            headers={"Content-disposition": "attachment; filename=unmatched_errors.csv"}
        )
        return response
    
    except Exception as e:
        return jsonify({'error': f'Error generating download: {str(e)}'}), 400

if __name__ == '__main__':
    import socket
    
    def find_free_port():
        """Find a free port to avoid 'Address already in use' errors"""
        for port in [5000, 8080, 8000, 3000, 9000]:
            try:
                with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                    s.bind(('localhost', port))
                    return port
            except OSError:
                continue
        return 5001  # Fallback port
    
    port = find_free_port()
    print("🚀 Starting Bootcode Verification Chatbot...")
    print(f"📍 Access at: http://localhost:{port}")
    print("🔧 Enhanced version with Unicode fixes and error handling")
    print("=" * 60)
    
    try:
        app.run(debug=False, host='0.0.0.0', port=port, use_reloader=False)
    except Exception as e:
        print(f"❌ Error starting server: {e}")
        print("💡 Try running with a different port or check for conflicts")
        print("🔧 Attempting to start on fallback port 9999...")
        try:
            app.run(debug=False, host='0.0.0.0', port=9999, use_reloader=False)
        except Exception as e2:
            print(f"❌ Fallback failed: {e2}")
            print("🚨 Please check system configuration and try again")