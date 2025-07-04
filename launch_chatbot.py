#!/usr/bin/env python3
"""
Simplified Launcher for Bootcode Verification Chatbot
Fixes hanging and startup issues
"""

import sys
import os

# Disable all logging that might cause conflicts
import logging
logging.getLogger().setLevel(logging.ERROR)

# Import Flask components
from flask import Flask, render_template, request, jsonify, session
import csv
import os
from werkzeug.utils import secure_filename
import uuid
from difflib import SequenceMatcher
import re
from collections import defaultdict
from datetime import datetime
import json
import io

# Recreate the app without the problematic logging setup
app = Flask(__name__)
app.secret_key = 'bootcode_verification_secret_key'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

# Create uploads directory if it doesn't exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Simple data storage
uploaded_data = {}
conversation_context = {}
unmatched_errors = []

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'csv', 'xls', 'xlsx'}

def similarity(a, b):
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()

def read_file_data(file):
    try:
        if file.filename.lower().endswith('.csv'):
            content = file.read().decode('utf-8')
            csv_reader = csv.DictReader(io.StringIO(content))
            data = list(csv_reader)
            columns = csv_reader.fieldnames
        else:
            from openpyxl import load_workbook
            workbook = load_workbook(file)
            sheet = workbook.active
            columns = [cell.value for cell in sheet[1]]
            data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):
                    row_dict = {}
                    for i, value in enumerate(row):
                        if i < len(columns) and columns[i]:
                            row_dict[columns[i]] = value if value is not None else ''
                    data.append(row_dict)
        return data, columns
    except Exception as e:
        raise Exception(f"Error reading file: {str(e)}")

def find_best_matches(error_message, data, threshold=0.6):
    if not data or len(data) == 0:
        return []
    
    columns = list(data[0].keys())
    error_col = columns[0]
    matches = []
    error_message_lower = error_message.lower()
    
    for row in data:
        stored_error = str(row.get(error_col, '')).lower()
        
        if error_message_lower in stored_error or stored_error in error_message_lower:
            similarity_score = 1.0
        else:
            similarity_score = similarity(error_message, str(row.get(error_col, '')))
        
        if similarity_score >= threshold:
            fixes = []
            priority = "Medium"
            
            for col_idx, col_name in enumerate(columns[1:], 1):
                if col_name in row:
                    col_value = str(row[col_name]).strip()
                    
                    if col_value and col_value.lower() not in ['nan', 'none', '', 'null']:
                        if col_name.lower() in ['priority', 'priority_level', 'urgency']:
                            priority = col_value
                        elif col_idx == 1:
                            fixes.insert(0, {'type': 'Primary', 'content': col_value})
                        elif col_idx == 2:
                            fixes.append({'type': 'Alternative', 'content': col_value})
                        elif col_idx == 3:
                            fixes.append({'type': 'Additional', 'content': col_value})
                        else:
                            fixes.append({'type': f'Option {col_idx-1}', 'content': col_value})
            
            if not fixes:
                fixes = [{'type': 'Primary', 'content': 'No specific fix provided'}]
            
            matches.append({
                'error': row.get(error_col, ''),
                'fixes': fixes,
                'priority': priority,
                'similarity': similarity_score
            })
    
    priority_order = {'High': 3, 'Medium': 2, 'Low': 1, 'Critical': 4, 'Urgent': 4}
    matches.sort(key=lambda x: (priority_order.get(x['priority'], 2), x['similarity']), reverse=True)
    return matches[:5]

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file selected'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if file and allowed_file(file.filename):
        try:
            if 'session_id' not in session:
                session['session_id'] = str(uuid.uuid4())
            
            session_id = session['session_id']
            data, columns = read_file_data(file)
            
            if len(columns) < 2:
                return jsonify({'error': 'File must have at least 2 columns'}), 400
            
            uploaded_data[session_id] = data
            
            return jsonify({
                'success': True, 
                'message': f'File uploaded successfully! Found {len(data)} error records.',
                'columns': list(columns),
                'sample_data': data[:3] if len(data) >= 3 else data
            })
            
        except Exception as e:
            return jsonify({'error': f'Error processing file: {str(e)}'}), 400
    
    return jsonify({'error': 'Invalid file type. Please upload CSV, XLS, or XLSX files.'}), 400

@app.route('/chat', methods=['POST'])
def chat():
    request_data = request.get_json()
    message = request_data.get('message', '').strip()
    
    if not message:
        return jsonify({'error': 'Please enter an error message'}), 400
    
    session_id = session.get('session_id')
    if not session_id or session_id not in uploaded_data:
        return jsonify({'error': 'Please upload a file first'}), 400
    
    data = uploaded_data[session_id]
    matches = find_best_matches(message, data)
    
    if not matches:
        response = {
            'message': 'No matching errors found in the uploaded data.',
            'suggestions': [
                'Try rephrasing your error message',
                'Check for typos',
                'Upload a more comprehensive error database'
            ]
        }
    else:
        if matches[0]['similarity'] >= 0.9:
            response = {
                'message': 'Found exact match! Here are the recommended solutions:',
                'exact_match': True,
                'matches': [matches[0]]
            }
        else:
            response = {
                'message': f'Found {len(matches)} similar error(s). Here are the closest matches:',
                'exact_match': False,
                'matches': matches
            }
    
    return jsonify(response)

@app.route('/clear', methods=['POST'])
def clear_session():
    session_id = session.get('session_id')
    if session_id and session_id in uploaded_data:
        del uploaded_data[session_id]
    session.clear()
    return jsonify({'success': True, 'message': 'Session cleared successfully'})

if __name__ == '__main__':
    print("🚀 Starting Bootcode Verification Chatbot...")
    print("📍 Access at: http://localhost:8080")
    print("🔧 Simplified launcher - no debug mode conflicts")
    print("=" * 50)
    
    # Simple startup without problematic logging
    app.run(host='0.0.0.0', port=8080, debug=False, use_reloader=False)